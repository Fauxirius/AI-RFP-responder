import PyPDF2
import docx
import io

import openpyxl

def extract_text_from_pdf(file_stream):
    """
    Extracts text from a PDF file stream.
    """
    try:
        pdf_reader = PyPDF2.PdfReader(file_stream)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return clean_text(text)
    except Exception as e:
        return f"Error reading PDF: {str(e)}"

def extract_text_from_docx(file_stream):
    """
    Extracts text from a DOCX file stream.
    """
    try:
        doc = docx.Document(file_stream)
        text = ""
        for para in doc.paragraphs:
            text += para.text + "\n"
        return clean_text(text)
    except Exception as e:
        return f"Error reading DOCX: {str(e)}"

def extract_text_from_excel(file_stream):
    """
    Extracts text from an Excel file stream, flattening it to markdown-like text.
    """
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"Sheet: {sheet}\n"
            for row in ws.iter_rows(values_only=True):
                # Filter out None values and join with |
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        return clean_text(text)
    except Exception as e:
        return f"Error reading Excel: {str(e)}"

def load_knowledge_base(directory_path):
    """
    Loads all supported files from a directory and concatenates their text.
    """
    kb_text = ""
    if not os.path.exists(directory_path):
        return ""
        
    for filename in os.listdir(directory_path):
        filepath = os.path.join(directory_path, filename)
        if os.path.isfile(filepath):
            file_text = ""
            with open(filepath, 'rb') as f:
                if filename.endswith('.pdf'):
                    file_text = extract_text_from_pdf(f)
                elif filename.endswith('.docx'):
                    file_text = extract_text_from_docx(f)
                elif filename.endswith('.xlsx'):
                    file_text = extract_text_from_excel(f)
            
            if file_text:
                kb_text += f"--- SOURCE DOCUMENT: {filename} ---\n{file_text}\n\n"
                
    return kb_text

def clean_text(text):
    """
    Basic text cleaning to remove excessive whitespace.
    """
    if not text:
        return ""
    # Replace multiple newlines with single newline
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return "\n".join(lines)
